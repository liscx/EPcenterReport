"""
标桥收益统计 — 数据导出模块
前提：已通过 login.py 完成登录并进入运营中心视角页面

步骤：
1. 点击侧边栏"标桥"一级菜单，展开子菜单
2. 点击"标桥收益统计"二级菜单，进入对应页面
3. 在筛选条件中：取消全选，只勾选"投标工具（含质量管理系统）"
4. 设置统计时间为上月（月报月份）
5. 点击搜索
"""

from datetime import datetime
import os
import time

import openpyxl
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException


def click_biaoqiao_sidebar(driver):
    """
    在运营中心视角页面，通过侧边栏导航到"标桥收益统计"：
    1. 点击一级菜单"标桥"（展开子菜单）
    2. 点击二级菜单"标桥收益统计"
    """
    print("正在导航到 标桥收益统计...")

    # --- 第一步：点击一级菜单"标桥" ---
    try:
        biaoqiao_menu = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable(
                (By.CSS_SELECTOR, 'li.level-1 a[data-id="00180005"]')
            )
        )
        biaoqiao_menu.click()
        print("  → 已点击 标桥 一级菜单")
        time.sleep(1)
    except TimeoutException:
        print("  → 未找到 标桥 菜单项！")
        raise

    # --- 第二步：点击二级菜单"标桥收益统计" ---
    try:
        stats_menu = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable(
                (By.CSS_SELECTOR, 'li.level-2 a[data-id="001800050001"]')
            )
        )
        stats_menu.click()
        print("  → 已点击 标桥收益统计")
        time.sleep(2)
    except TimeoutException:
        print("  → 未找到 标桥收益统计 菜单项！")
        raise


def _get_report_month():
    """
    根据当前日期计算月报月份。
    当前 6 月 → 月报为 5 月 → 返回 '2026-05'
    当前 7 月 → 月报为 6 月 → 返回 '2026-06'
    """
    today = datetime.today()
    year = today.year
    month = today.month - 1  # 上月
    if month == 0:  # 1 月的情况：上月为去年 12 月
        year -= 1
        month = 12
    return f"{year}-{month:02d}"


def set_filter_conditions(driver):
    """
    在标桥收益统计页面设置筛选条件：
    1. 切换到内容 iframe
    2. 取消全选，只勾选"投标工具（含质量管理系统）"
    3. 设置统计时间（开始/结束）为月报月份
    4. 点击搜索按钮
    """
    report_month = _get_report_month()
    print(f"正在设置筛选条件，月报月份: {report_month}...")

    # 切换到内容 iframe（标桥收益统计页面加载在 iframe 中）
    _switch_to_content_iframe(driver)

    # --- 第一步：设置 BP 子产品复选框 ---
    # miniui checkboxlist: #bpType
    # 选项：0=清标工具, 1=建采通, 2=投标工具（含质量管理系统）, 3=标讯, 4=智能客服, 5=采购学院
    # 先取消全选，再单独勾选 index=0（清标工具）
    driver.execute_script("""
        var bpType = mini.get('bpType');
        bpType.deselectAll();
        bpType.select(0);
    """)
    print("  → 已勾选 清标工具")

    # --- 第二步：设置统计时间 ---
    # miniui monthpicker: #startDate, #endDate
    # 使用 miniui API 直接设置值
    driver.execute_script(f"""
        var startDate = mini.get('startDate');
        var endDate = mini.get('endDate');
        startDate.setValue('{report_month}');
        endDate.setValue('{report_month}');
    """)
    print(f"  → 已设置统计时间: {report_month}")

    time.sleep(1)

    # --- 第三步：点击搜索按钮 ---
    search_btn = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.CSS_SELECTOR, "a.cond-srh-btn"))
    )
    search_btn.click()
    print("  → 已点击搜索，等待数据加载...")
    time.sleep(3)


def _switch_to_content_iframe(driver):
    """
    切换到标桥收益统计页面的内容 iframe。
    iframe id 为 tab-content-001800050001（对应菜单 data-id="001800050001"）。
    """
    # 先切回默认内容
    driver.switch_to.default_content()
    time.sleep(1)

    # 标桥收益统计的 iframe id 固定为 tab-content-001800050001
    iframe = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.ID, "tab-content-001800050001"))
    )
    driver.switch_to.frame(iframe)
    print("  → 已切换到标桥收益统计 iframe")


def _switch_to_top_popup_iframe(driver):
    """
    识别当前页面（或当前 iframe）中所有的弹窗 iframe，并切入到最顶层（最新打开）的那个。
    """
    try:
        # MiniUI 弹窗通常是 .mini-iframe 或 mini-window 下的 iframe
        popups = driver.find_elements(By.CSS_SELECTOR, "iframe.mini-iframe, .mini-window iframe")
        if popups:
            # 切换到最后一个找到的 iframe（通常是 z-index 最高的）
            driver.switch_to.frame(popups[-1])
            return True
    except:
        pass
    return False


# ============================================================
# 数据采集与导出
# ============================================================

def _get_report_month_str():
    """返回用于文件路径的月份字符串，如 '202605'"""
    return _get_report_month().replace("-", "")


def click_view_button_in_results(driver):
    """
    在搜索结果表格中，点击第一行的"查看"按钮（icon-search），
    打开产品收益弹框。
    """
    print("正在点击搜索结果中的 查看 按钮...")
    _switch_to_content_iframe(driver)
    view_btn = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable(
            (By.CSS_SELECTOR, "i.action-icon.icon-search")
        )
    )
    view_btn.click()
    print("  → 已点击 查看，等待弹框加载...")
    time.sleep(3)

    # 弹窗通常在当前内容 iframe 内部，需要切入进去
    if _switch_to_top_popup_iframe(driver):
        print("  → 已进入 产品收益 弹窗内部")


def _parse_revenue_table(driver):
    """
    从产品收益弹框中"02 收益组成"的 fui-acc-item 区域，
    直接解析表格数据并返回。
    """
    time.sleep(2)

    # 通过查找包含特定标题文本的面板来定位"02 收益组成"
    acc_items = driver.find_elements(By.CSS_SELECTOR, ".fui-acc-item")
    target = None
    for item in acc_items:
        try:
            hd = item.find_element(By.CSS_SELECTOR, ".fui-acc-hd")
            if "02 收益组成" in hd.text:
                target = item.find_element(By.CSS_SELECTOR, ".fui-acc-bd")
                print(f"  → 精确匹配到面板: {hd.text.strip()}")
                break
        except:
            continue

    # 如果没按标题搜到，退而求其次寻找包含 grid 的面板
    if not target:
        acc_bodies = driver.find_elements(By.CSS_SELECTOR, ".fui-acc-bd")
        for acc_bd in acc_bodies:
            headers = acc_bd.find_elements(By.CSS_SELECTOR, ".mini-grid-columns-view .mini-grid-headerCell-inner")
            if headers:
                target = acc_bd
                break

    if not target:
        return []

    # 获取表头
    headers = []
    header_cells = target.find_elements(By.CSS_SELECTOR, ".mini-grid-columns-view .mini-grid-headerCell-inner")
    for cell in header_cells:
        text = cell.text.strip()
        if text:
            headers.append(text)

    # 获取数据行
    rows = []
    row_elements = target.find_elements(By.CSS_SELECTOR, ".mini-grid-rows-view .mini-grid-row")
    for row_el in row_elements:
        cells = row_el.find_elements(By.CSS_SELECTOR, ".mini-grid-cell-inner")
        row_data = {}
        has_value = False
        for i, cell in enumerate(cells):
            if i < len(headers):
                val = (cell.get_attribute("title") or cell.text or "").strip()
                row_data[headers[i]] = val
                if val:
                    has_value = True
        if has_value:
            rows.append(row_data)

    return rows


def _close_top_window(driver):
    """
    关闭最顶层的 miniui 弹框窗口。
    """
    try:
        # 获取所有可见的关闭按钮
        close_btns = driver.find_elements(By.CSS_SELECTOR,
                                          ".mini-window:not([style*='display: none']) .mini-tools-close")
        if not close_btns:
            # 如果当前找不到，可能是在 iframe 内部，尝试切回父层找
            driver.switch_to.parent_frame()
            close_btns = driver.find_elements(By.CSS_SELECTOR,
                                              ".mini-window:not([style*='display: none']) .mini-tools-close")

        if close_btns:
            # 使用 JS 点击最顶层的关闭按钮，避开可能的拦截
            driver.execute_script("arguments[0].click();", close_btns[-1])
            time.sleep(1.5)
            print("  → 已关闭弹窗")
            return True
    except Exception as e:
        print(f"  → 关闭弹窗失败: {e}")
    return False


def export_revenue_data(driver, output_dir):
    """
    完整数据采集流程：
    1. 设置筛选条件为清标工具
    2. 点击搜索结果的"查看"按钮，打开产品收益弹框
    3. 直接从"02 收益组成"表格获取数据并保存
    """
    os.makedirs(output_dir, exist_ok=True)
    output_file = os.path.join(output_dir, "营收平台标桥收益数据.xlsx")

    # 第一步：设置筛选条件为清标工具
    set_filter_conditions(driver)

    # 第二步：打开产品收益主弹框
    click_view_button_in_results(driver)

    # 第二步：直接解析"03 收益组成"表格
    print("正在解析 03 收益组成 表格...")
    rows = _parse_revenue_table(driver)
    print(f"  → 获取到 {len(rows)} 条数据")

    if not rows:
        print("  → 没有数据，跳过导出")
        _close_top_window(driver)
        return output_file

    # 第三步：写入 Excel（追加模式）
    if os.path.exists(output_file):
        wb = openpyxl.load_workbook(output_file)
        # 如果"清标工具" sheet 已存在，删除后重建（避免重复数据）
        if "清标工具" in wb.sheetnames:
            del wb["清标工具"]
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

    _write_to_sheet(wb, "清标工具", rows)

    # 关闭弹框
    _close_top_window(driver)

    wb.save(output_file)
    print(f"\n✅ 数据已保存到: {output_file}")
    return output_file


def _write_to_sheet(wb, sheet_name, rows):
    """
    将数据写入 Excel 的指定 sheet。
    使用显式行索引确保不出现空行。
    """
    safe_name = sheet_name.replace("/", "").replace("\\", "").replace("*", "")[:31]
    if safe_name in wb.sheetnames:
        ws = wb[safe_name]
    else:
        ws = wb.create_sheet(title=safe_name)

    # 确定数据写入的具体物理行号
    # 统计 A 列中真正有值的行数，以避开 openpyxl 的 max_row 虚假位移
    real_data_rows = 0
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1, values_only=True):
        if row[0] is not None:
            real_data_rows += 1

    if real_data_rows == 0:
        # 完全没有数据，强制在第 1 行写入表头
        cols = ["序", "分公司", "项目名称", "类型", "收益（元）", "应收（元）", "来源"]
        for i, val in enumerate(cols, 1):
            ws.cell(row=1, column=i, value=val)
        write_row = 2
    else:
        # 已经有数据了，从下一行开始追加
        write_row = real_data_rows + 1

    # 显式按行号写入数据，不使用可能导致空行的 append
    # 注意：源数据中列顺序是 分公司、项目名称、类型、序、收益（元）、应收（元）、来源
    # 写入时调整为：序、分公司、项目名称、类型、收益（元）、应收（元）、来源
    for data in rows:
        ws.cell(row=write_row, column=1, value=data.get("序", ""))
        ws.cell(row=write_row, column=2, value=data.get("分公司", ""))
        ws.cell(row=write_row, column=3, value=data.get("项目名称", ""))
        ws.cell(row=write_row, column=4, value=data.get("类型", ""))
        ws.cell(row=write_row, column=5, value=data.get("收益（元）", ""))
        ws.cell(row=write_row, column=6, value=data.get("应收（元）", ""))
        ws.cell(row=write_row, column=7, value=data.get("来源", ""))
        write_row += 1


if __name__ == "__main__":
    pass
