"""
标桥收益统计 — 数据导出模块
前提：已通过 login.py 完成登录并进入运营中心视角页面

步骤：
1. 点击侧边栏"标桥"一级菜单，展开子菜单
2. 点击"标桥收益统计"二级菜单，进入对应页面
3. 在筛选条件中：取消全选，只勾选"投标工具（含质量管理系统）"
4. 设置统计时间为上月（月报月份）
5. 点击搜索
"""

from datetime import datetime
import os
import time

import openpyxl
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException


def click_biaoqiao_sidebar(driver):
    """
    在运营中心视角页面，通过侧边栏导航到"标桥收益统计"：
    1. 点击一级菜单"标桥"（展开子菜单）
    2. 点击二级菜单"标桥收益统计"
    """
    print("正在导航到 标桥收益统计...")

    # 先切回默认内容（避免之前在 iframe 中）
    driver.switch_to.default_content()
    time.sleep(1)

    # --- 第一步：点击一级菜单"标桥" ---
    try:
        biaoqiao_menu = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable(
                (By.CSS_SELECTOR, 'li.level-1 a[data-id="00180005"]')
            )
        )
        biaoqiao_menu.click()
        print("  → 已点击 标桥 一级菜单")
        time.sleep(1)
    except TimeoutException:
        print("  → 未找到 标桥 菜单项！")
        raise

    # --- 第二步：点击二级菜单"标桥收益统计" ---
    try:
        stats_menu = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable(
                (By.CSS_SELECTOR, 'li.level-2 a[data-id="001800050001"]')
            )
        )
        stats_menu.click()
        print("  → 已点击 标桥收益统计")
        time.sleep(2)
    except TimeoutException:
        print("  → 未找到 标桥收益统计 菜单项！")
        raise


def _get_report_month(year=None, month=None):
    """
    根据当前日期计算月报月份。
    当前 6 月 → 月报为 5 月 → 返回 '2026-05'
    当前 7 月 → 月报为 6 月 → 返回 '2026-06'

    Args:
        year: 指定年份，如 2025
        month: 指定月份，如 5
    """
    if year is not None and month is not None:
        return f"{year}-{month:02d}"

    today = datetime.today()
    year = today.year
    month = today.month - 1  # 上月
    if month == 0:  # 1 月的情况：上月为去年 12 月
        year -= 1
        month = 12
    return f"{year}-{month:02d}"


def set_filter_conditions(driver, year=None, month=None):
    """
    在标桥收益统计页面设置筛选条件：
    1. 切换到内容 iframe
    2. 取消全选，只勾选"投标工具（含质量管理系统）"
    3. 设置统计时间（开始/结束）为月报月份
    4. 点击搜索按钮
    """
    report_month = _get_report_month(year, month)
    print(f"正在设置筛选条件，月报月份: {report_month}...")

    # 切换到内容 iframe（标桥收益统计页面加载在 iframe 中）
    _switch_to_content_iframe(driver)

    # --- 第一步：设置 BP 子产品复选框 ---
    # miniui checkboxlist: #bpType
    # 选项：0=清标工具, 1=建采通, 2=投标工具（含质量管理系统）, 3=标讯, 4=智能客服, 5=采购学院
    # 先取消全选，再单独勾选 index=2
    driver.execute_script("""
        var bpType = mini.get('bpType');
        bpType.deselectAll();
        bpType.select(2);
    """)
    print("  → 已勾选 投标工具（含质量管理系统）")

    # --- 第二步：设置统计时间 ---
    # miniui monthpicker: #startDate, #endDate
    # 使用 miniui API 直接设置值
    driver.execute_script(f"""
        var startDate = mini.get('startDate');
        var endDate = mini.get('endDate');
        startDate.setValue('{report_month}');
        endDate.setValue('{report_month}');
    """)
    print(f"  → 已设置统计时间: {report_month}")

    time.sleep(1)

    # --- 第三步：点击搜索按钮 ---
    search_btn = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.CSS_SELECTOR, "a.cond-srh-btn"))
    )
    search_btn.click()
    print("  → 已点击搜索，等待数据加载...")
    time.sleep(3)


def _switch_to_content_iframe(driver):
    """
    切换到标桥收益统计页面的内容 iframe。
    iframe id 为 tab-content-001800050001（对应菜单 data-id="001800050001"）。
    """
    # 先切回默认内容
    driver.switch_to.default_content()
    time.sleep(1)

    # 标桥收益统计的 iframe id 固定为 tab-content-001800050001
    iframe = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.ID, "tab-content-001800050001"))
    )
    driver.switch_to.frame(iframe)
    print("  → 已切换到标桥收益统计 iframe")


def _switch_to_top_popup_iframe(driver):
    """
    识别当前页面（或当前 iframe）中所有的弹窗 iframe，并切入到最顶层（最新打开）的那个。
    """
    try:
        # MiniUI 弹窗通常是 .mini-iframe 或 mini-window 下的 iframe
        popups = driver.find_elements(By.CSS_SELECTOR, "iframe.mini-iframe, .mini-window iframe")
        if popups:
            # 切换到最后一个找到的 iframe（通常是 z-index 最高的）
            driver.switch_to.frame(popups[-1])
            return True
    except:
        pass
    return False


# ============================================================
# 数据采集与导出
# ============================================================

def _get_report_month_str():
    """返回用于文件路径的月份字符串，如 '202605'"""
    return _get_report_month().replace("-", "")


def click_view_button_in_results(driver):
    """
    在搜索结果表格中，点击第一行的"查看"按钮（icon-search），
    打开产品收益弹框。
    """
    print("正在点击搜索结果中的 查看 按钮...")
    _switch_to_content_iframe(driver)
    view_btn = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable(
            (By.CSS_SELECTOR, "i.action-icon.icon-search")
        )
    )
    view_btn.click()
    print("  → 已点击 查看，等待弹框加载...")
    time.sleep(3)

    # 弹窗通常在当前内容 iframe 内部，需要切入进去
    if _switch_to_top_popup_iframe(driver):
        print("  → 已进入 产品收益 弹窗内部")


def _get_sub_products_info(driver):
    """
    从产品收益弹框中"03 收益组成"的 fui-acc-bd 区域，
    获取所有子产品信息：名称 + 对应的查看按钮。
    返回 list[dict]，每项: {name, button_element}
    """
    time.sleep(2)

    # 找到所有 fui-acc-bd（accordion body），定位"03 收益组成"那个
    acc_bodies = driver.find_elements(By.CSS_SELECTOR, ".fui-acc-bd")
    for acc_bd in acc_bodies:
        buttons = acc_bd.find_elements(By.CSS_SELECTOR, "i.action-icon.icon-search")
        if not buttons:
            continue

        # 从每个按钮所在的行<tr>中提取子产品名称（第2个cell）
        results = []
        for btn in buttons:
            try:
                row = btn.find_element(By.XPATH, "./ancestor::tr")
                # MiniUI 的表格单元格内容可能在 .mini-grid-cell-inner 中
                cells = row.find_elements(By.CSS_SELECTOR, ".mini-grid-cell-inner")
                name = ""
                if len(cells) >= 2:
                    # 尝试从 title 属性或 text 获取名称
                    name = (cells[1].get_attribute("title") or cells[1].text or "").strip()
                results.append({"name": name, "button": btn})
            except Exception as e:
                print(f"  → 解析某一行子产品时出错: {e}")
                continue
        return results

    return []


def _parse_detail_grid_html(driver):
    """
    从子产品详情弹框中，精确定位名为"02 收益组成"的面板并解析其表格数据。
    """
    time.sleep(1.5)

    # 通过查找包含特定标题文本的面板来定位
    acc_items = driver.find_elements(By.CSS_SELECTOR, ".fui-acc-item")
    target_panel = None
    
    for item in acc_items:
        try:
            hd = item.find_element(By.CSS_SELECTOR, ".fui-acc-hd")
            if "02 收益组成" in hd.text:
                target_panel = item
                print(f"    → 精确匹配到面板: {hd.text.strip()}")
                break
        except:
            continue

    # 如果没按标题搜到，退而求其次寻找包含 grid 的面板
    if not target_panel:
        acc_bodies = driver.find_elements(By.CSS_SELECTOR, ".fui-acc-bd")
        if not acc_bodies:
            return []
        # fallback: 假设第2个或最后一个是目标
        target = acc_bodies[1] if len(acc_bodies) > 1 else acc_bodies[-1]
    else:
        target = target_panel.find_element(By.CSS_SELECTOR, ".fui-acc-bd")

    # 获取表头
    headers = []
    header_cells = target.find_elements(
        By.CSS_SELECTOR, ".mini-grid-columns-view .mini-grid-headerCell-inner"
    )
    for cell in header_cells:
        text = cell.text.strip()
        if text:
            headers.append(text)

    # 获取数据行
    rows = []
    row_elements = target.find_elements(
        By.CSS_SELECTOR, ".mini-grid-rows-view .mini-grid-row"
    )
    for row_el in row_elements:
        cells = row_el.find_elements(By.CSS_SELECTOR, ".mini-grid-cell-inner")
        row_data = {}
        has_value = False
        for i, cell in enumerate(cells):
            if i < len(headers):
                val = (cell.get_attribute("title") or cell.text or "").strip()
                row_data[headers[i]] = val
                if val:
                    has_value = True
        if has_value:
            rows.append(row_data)

    return rows


def _close_top_window(driver):
    """
    关闭最顶层的 miniui 弹框窗口。
    """
    try:
        # 获取所有可见的关闭按钮
        close_btns = driver.find_elements(By.CSS_SELECTOR, ".mini-window:not([style*='display: none']) .mini-tools-close")
        if not close_btns:
            # 如果当前找不到，可能是在 iframe 内部，尝试切回父层找
            driver.switch_to.parent_frame()
            close_btns = driver.find_elements(By.CSS_SELECTOR, ".mini-window:not([style*='display: none']) .mini-tools-close")
            
        if close_btns:
            # 使用 JS 点击最顶层的关闭按钮，避开可能的拦截
            driver.execute_script("arguments[0].click();", close_btns[-1])
            time.sleep(1.5)
            print("  → 已关闭弹窗")
            return True
    except Exception as e:
        print(f"  → 关闭弹窗失败: {e}")
    return False


def export_revenue_data(driver, output_dir, year=None, month=None):
    """
    完整数据采集流程：
    1. 点击搜索结果的"查看"按钮，打开产品收益弹框
    2. 从"03 收益组成"表格中获取所有子产品信息
    3. 逐个点击子产品的"查看"按钮，打开详情弹框
    4. 解析详情数据并保存
    """
    month_str = _get_report_month(year, month).replace("-", "")
    os.makedirs(output_dir, exist_ok=True)
    output_file = os.path.join(output_dir, "营收平台标桥收益数据.xlsx")

    # 第一步：打开产品收益主弹框
    click_view_button_in_results(driver)

    # 第二步：获取子产品列表（此时应在第一个弹窗 iframe 内）
    print("正在解析 03 收益组成 表格...")
    sub_products = _get_sub_products_info(driver)
    print(f"  → 找到 {len(sub_products)} 个子产品")

    if not sub_products:
        print("  → 没有子产品数据，跳过导出")
        _close_top_window(driver)
        return output_file

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # 第三步：遍历子产品
    for idx in range(len(sub_products)):
        # 每次循环开始，确保回到第一个弹窗的 iframe 环境
        _switch_to_content_iframe(driver)
        _switch_to_top_popup_iframe(driver)
        
        # 重新获取最新的按钮列表，防止 StaleElementReferenceException
        current_list = _get_sub_products_info(driver)
        if idx >= len(current_list):
            break
            
        product_name = current_list[idx]["name"] or f"子产品{idx+1}"
        target_btn = current_list[idx]["button"]
        
        print(f"  [{idx+1}/{len(sub_products)}] 正在处理: {product_name}")

        try:
            # 使用 JS 点击，解决 "element click intercepted"
            driver.execute_script("arguments[0].click();", target_btn)
            print(f"    → 已点击查看按钮，等待详情弹窗加载...")

            # 无限等待二级详情弹窗出现（其他营运项目加载较慢，用10秒间隔）
            check_interval = 10 if product_name == "其他营运项目" else 2
            while True:
                popups = driver.find_elements(By.CSS_SELECTOR, "iframe.mini-iframe, .mini-window iframe")
                if popups:
                    print(f"    → 详情弹窗已加载")
                    time.sleep(3)  # 等待数据加载
                    break
                time.sleep(check_interval)

            # 进入二级详情弹窗
            if _switch_to_top_popup_iframe(driver):
                # 其他营运项目加载较慢，先等1分钟，数据为0则每分钟重试，最多5分钟
                if product_name == "其他营运项目":
                    print(f"    → 其他营运项目：强制等待1分钟...")
                    time.sleep(60)
                    rows = _parse_detail_grid_html(driver)
                    waited = 60
                    while len(rows) == 0 and waited < 300:
                        print(f"    → 数据量为0，继续等待1分钟...（已等待{waited}秒）")
                        time.sleep(60)
                        waited += 60
                        rows = _parse_detail_grid_html(driver)
                    if len(rows) == 0:
                        print(f"    → [警告] 已等待5分钟，数据量仍为0")
                else:
                    rows = _parse_detail_grid_html(driver)

                print(f"    → 获取到 {len(rows)} 条数据")
                _write_to_sheet(wb, product_name, rows)

                # 关闭二级弹窗
                # 注意：此时在二级 iframe 内，_close_top_window 会尝试切回一级 iframe 执行关闭
                _close_top_window(driver)
            else:
                print(f"    → [错误] 无法进入 {product_name} 的详情窗")

        except Exception as e:
            print(f"    → 处理 {product_name} 异常: {e}")
            # 异常时尝试清理并寻找回到主环境的路径
            _switch_to_content_iframe(driver)
            _switch_to_top_popup_iframe(driver)
            _close_top_window(driver)
            continue

    # 处理完毕，关闭主弹框
    _switch_to_content_iframe(driver)
    _close_top_window(driver)

    wb.save(output_file)
    print(f"\n✅ 数据已保存到: {output_file}")
    return output_file


def _write_to_sheet(wb, sheet_name, rows):
    """
    将数据写入 Excel 的指定 sheet。
    使用显式行索引确保不出现空行。
    """
    safe_name = sheet_name.replace("/", "").replace("\\", "").replace("*", "")[:31]
    if safe_name in wb.sheetnames:
        ws = wb[safe_name]
    else:
        ws = wb.create_sheet(title=safe_name)

    # 确定数据写入的具体物理行号
    # 统计 A 列中真正有值的行数，以避开 openpyxl 的 max_row 虚假位移
    real_data_rows = 0
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1, values_only=True):
        if row[0] is not None:
            real_data_rows += 1
    
    if real_data_rows == 0:
        # 完全没有数据，强制在第 1 行写入表头
        cols = ["序", "分公司", "项目名称", "子产品", "订单金额（元）", "收益金额（元）", "来源"]
        for i, val in enumerate(cols, 1):
            ws.cell(row=1, column=i, value=val)
        write_row = 2
    else:
        # 已经有数据了，从下一行开始追加
        write_row = real_data_rows + 1

    # 显式按行号写入数据，不使用可能导致空行的 append
    for data in rows:
        ws.cell(row=write_row, column=1, value=data.get("序", ""))
        ws.cell(row=write_row, column=2, value=data.get("分公司", ""))
        ws.cell(row=write_row, column=3, value=data.get("项目名称", ""))
        ws.cell(row=write_row, column=4, value=data.get("子产品", ""))
        ws.cell(row=write_row, column=5, value=data.get("订单金额（元）", ""))
        ws.cell(row=write_row, column=6, value=data.get("收益金额（元）", ""))
        ws.cell(row=write_row, column=7, value=data.get("来源", ""))
        write_row += 1


if __name__ == "__main__":
    pass
